from google.oauth2.credentials import Credentials
from google_auth_oauthlib.flow import InstalledAppFlow
from google.auth.transport.requests import Request
from googleapiclient.discovery import build
from email.utils import parsedate_to_datetime
import os
import time
import requests
from dotenv import load_dotenv
from openpyxl import Workbook, load_workbook
import json
import pandas as pd


load_dotenv()

SCOPES = ['https://www.googleapis.com/auth/gmail.readonly']

class GmailService:
    def __init__(self):
        creds = None
        if os.path.exists('gmail_token.json'):
            creds = Credentials.from_authorized_user_file('gmail_token.json', SCOPES)
        if not creds or not creds.valid:
            if creds and creds.expired and creds.refresh_token:
                creds.refresh(Request())
            else:
                flow = InstalledAppFlow.from_client_secrets_file('credentials.json', SCOPES)
                creds = flow.run_local_server(port=0)
            with open('gmail_token.json', 'w') as token:
                token.write(creds.to_json()

                )
        self.service = build('gmail', 'v1', credentials=creds)
        self.important_emails = []

    def extract_job_related_emails_to_excel(self):
        try:
            last_run_file = "last_run.txt"
            if os.path.exists(last_run_file):
                with open(last_run_file, "r") as f:
                    last_run_time = int(f.read().strip())
            else:
                # Default: 3 days ago if no previous run
                last_run_time = int(time.time()) - 3 * 86400

            query = f"after:{last_run_time}"
            messages = []
            result = self.service.users().messages().list(
                userId='me',
                labelIds=['INBOX'],
                q=query
            ).execute()

            while 'messages' in result:
                messages.extend(result['messages'])
                if 'nextPageToken' in result:
                    result = self.service.users().messages().list(
                        userId='me',
                        labelIds=['INBOX'],
                        q=query,
                        pageToken=result['nextPageToken']
                    ).execute()
                else:
                    break

            structured_jobs = []
            count = 0
            for msg in messages:
                msg_data = self.service.users().messages().get(userId='me', id=msg['id']).execute()
                headers = msg_data['payload']['headers']
                subject = next((h['value'] for h in headers if h['name'] == 'Subject'), "No Subject")
                sender = next((h['value'] for h in headers if h['name'] == 'From'), "Unknown Sender")
                snippet = msg_data.get('snippet', '')
                date_str = next((h['value'] for h in headers if h['name'].lower() == 'date'), None)
                email_date = parsedate_to_datetime(date_str).strftime('%Y-%m-%d') if date_str else "Unknown"

                if is_job_email_by_keywords(subject, snippet):
                    count += 1
                    print(f'job email found:{count}', subject)
                    job = extract_structured_job_data(subject, sender, snippet, email_date)
                    if job:
                        structured_jobs.append(job)

            save_structured_jobs_to_excel(structured_jobs)

            # Update last run timestamp
            with open(last_run_file, "w") as f:
                f.write(str(int(time.time())))

            return len(structured_jobs)

        except Exception as e:
            print("❌ Job email extraction failed:", e)
            return 0


# -------- Keywords Filter --------
def is_job_email_by_keywords(subject, snippet):
    keywords = [
        "application", "applied", "interview", "opportunity", "position", "opening",
        "recruiter", "job", "career", "hiring", "role", "thank you for applying"
    ]
    text = f"{subject} {snippet}".lower()
    # print("🔍 Checking keywords in email:", text)
    return any(kw in text for kw in keywords)

# -------- LLM Extractor --------

import re
import json

def extract_structured_job_data(subject, sender, snippet, email_date):
    prompt = f"""
You are a job email parser. Extract job info ONLY from application confirmation/status update emails.

Respond with a JSON object in **this exact format**:
{{
  "company": "Company Name",
  "role": "Job Title",
  "status": "Applied / Interview Scheduled / Rejected / Offer",
  "date": "{email_date}"
}}

⚠️ DO NOT return markdown formatting like triple backticks or explanations.
If the email is not a job status email, return exactly "No".

Subject: {subject}
Sender: {sender}
Content: {snippet}
"""

    headers = {
        "Authorization": f"Bearer {os.getenv('OPENROUTER_API_KEY')}",
        "Content-Type": "application/json",
    }

    data = {
        "model": "mistral/ministral-8b",
        "messages": [
            { "role": "system", "content": "You extract structured job info from emails." },
            { "role": "user", "content": prompt }
        ]
    }

    try:
        res = requests.post("https://openrouter.ai/api/v1/chat/completions", headers=headers, json=data)
        result = res.json()["choices"][0]["message"]["content"].strip()

        if result.lower() == "no":
            return None

        # Remove markdown formatting if present
        cleaned = re.sub(r"^```json|```$", "", result.strip(), flags=re.IGNORECASE).strip()

        # Try parsing the cleaned result
        try:
            job_data = json.loads(cleaned)
        except json.JSONDecodeError:
            print("⚠️ JSON parse error:", result)
            return None

        # Validate required fields
        required = ["company", "role", "status", "date"]
        job_data = {k: job_data.get(k, "").strip() for k in required}

        if all(job_data[k] for k in required):
            return job_data
        else:
            print("⚠️ Incomplete job data:", job_data)
            return None

    except Exception as e:
        print("⚠️ LLM Extraction Error:", e)
        return None



# -------- Excel Writer --------


def summarize_job_applications(file_path="job_emails.xlsx"):
    try:
        df = pd.read_excel(file_path)

        # Clean up and standardize status column
        df['Status'] = df['Status'].str.strip().str.lower()

        # Convert Date column to datetime
        df['Date'] = pd.to_datetime(df['Date'], errors='coerce')

        # Drop rows with invalid dates
        df = df.dropna(subset=['Date'])

        # Total counts
        total_applied = (df['Status'] == 'applied').sum()
        total_interviews = (df['Status'] == 'interview scheduled').sum()
        total_rejected = (df['Status'] == 'rejected').sum()
        total_offers = (df['Status'] == 'offer').sum()

        # Daily applications count with string keys
        daily_applications = (
            df[df['Status'] == 'applied']
            .groupby(df['Date'].dt.date)
            .size()
            .rename_axis("date")
            .reset_index(name="count")
        )

        # Convert to { "2025-05-16": 3, ... }
        daily_applications_dict = {
            str(row["date"]): int(row["count"]) for _, row in daily_applications.iterrows()
        }

        summary = {
            "total_applied": int(total_applied),
            "total_interview_scheduled": int(total_interviews),
            "total_rejected": int(total_rejected),
            "total_offers": int(total_offers),
            "daily_applications": daily_applications_dict
        }

        return summary
    except Exception as e:
        return {"error": str(e)}




def save_structured_jobs_to_excel(jobs, filename="job_emails.xlsx"):
    if os.path.exists(filename):
        wb = load_workbook(filename)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "Job Applications"
        ws.append(["Company", "Role", "Status", "Date"])  # write headers

    for job in jobs:
        if all(job.get(k) for k in ["company", "role", "status", "date"]):
            ws.append([job["company"], job["role"], job["status"], job["date"]])

    try:
        wb.save(filename)
        print(f"✅ Appended {len(jobs)} job emails to {filename}")
    except Exception as e:
        print("❌ Failed to save Excel file:", e)

